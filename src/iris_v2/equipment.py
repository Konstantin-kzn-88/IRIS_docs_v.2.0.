import json
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from iris_v2.substances import SubstanceError, SubstanceService


EXCEL_FILE_NAME = "equipment_data.xlsx"
JSON_FILE_NAME = "equipments.json"
ERROR_FILE_NAME = "equipment_import_errors.txt"
SHEET_NAME = "Equipment Data"
PIPELINE_TYPES = {0, 9}
PHASE_STATES = {"ж.ф.", "г.ф.", "г.ф.+ж.ф."}
HEADERS = (
    "source_id",
    "substance_id",
    "equipment_name",
    "equipment_type",
    "phase_state",
    "total_length_m",
    "equipment_count",
    "accident_section_length_m",
    "diameter_mm",
    "wall_thickness_mm",
    "volume_m3",
    "fill_fraction",
    "pressure_mpa",
    "spill_coefficient",
    "spill_area_m2",
    "substance_temperature_c",
    "shutdown_time_s",
    "evaporation_time_s",
    "hazard_component",
    "clutter_degree",
    "coord_type",
    "coordinate_x",
    "coordinate_y",
    "possible_dead",
    "possible_injured",
)


class EquipmentError(Exception):
    pass


@dataclass(frozen=True)
class EquipmentImportResult:
    count: int
    json_path: Path
    excel_path: Path


def _is_empty(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _text(value: Any, field: str, errors: list[str], row: int) -> str:
    result = "" if value is None else str(value).strip()
    if not result:
        errors.append(f"строка {row}, {field}: поле не заполнено")
    return result


def _number(
    value: Any,
    field: str,
    errors: list[str],
    row: int,
    *,
    required: bool = True,
) -> float | None:
    if _is_empty(value):
        if required:
            errors.append(f"строка {row}, {field}: поле не заполнено")
        return None
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        errors.append(f"строка {row}, {field}: ожидается число, получено {value!r}")
        return None
    return float(value)


def _integer(
    value: Any,
    field: str,
    errors: list[str],
    row: int,
    *,
    required: bool = True,
) -> int | None:
    number = _number(value, field, errors, row, required=required)
    if number is None:
        return None
    if not number.is_integer():
        errors.append(f"строка {row}, {field}: ожидается целое число")
        return None
    return int(number)


def _positive(value: float | None, field: str, errors: list[str], row: int) -> None:
    if value is not None and value <= 0:
        errors.append(f"строка {row}, {field}: значение должно быть больше нуля")


def _non_negative(
    value: float | int | None, field: str, errors: list[str], row: int
) -> None:
    if value is not None and value < 0:
        errors.append(f"строка {row}, {field}: значение не может быть отрицательным")


class EquipmentService:
    def ensure_template(self, project_directory: Path | str) -> Path:
        project = Path(project_directory)
        input_directory = project / "input"
        if not input_directory.is_dir():
            raise EquipmentError(f"Папка input не найдена: {input_directory}")
        destination = input_directory / EXCEL_FILE_NAME
        if not destination.exists():
            source = Path(__file__).parent / "data" / EXCEL_FILE_NAME
            if not source.is_file():
                raise EquipmentError("Встроенный шаблон equipment_data.xlsx не найден")
            shutil.copy2(source, destination)
        return destination

    def import_excel(
        self, project_directory: Path | str, excel_path: Path | str
    ) -> EquipmentImportResult:
        project = Path(project_directory)
        source = Path(excel_path)
        if not source.is_file():
            raise EquipmentError(f"Файл Excel не найден: {source}")

        try:
            project_substances = SubstanceService().load_project(project)
        except SubstanceError as exc:
            raise EquipmentError(str(exc)) from exc
        substance_ids = {
            item.get("id") for item in project_substances if isinstance(item.get("id"), int)
        }
        if not substance_ids:
            raise EquipmentError("Сначала выберите и сохраните вещества проекта")

        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise EquipmentError("Не установлена зависимость openpyxl") from exc

        try:
            workbook = load_workbook(source, read_only=True, data_only=True)
        except Exception as exc:
            raise EquipmentError(f"Не удалось открыть Excel: {source}") from exc
        try:
            if SHEET_NAME not in workbook.sheetnames:
                raise EquipmentError(f"В Excel отсутствует лист {SHEET_NAME!r}")
            sheet = workbook[SHEET_NAME]
            actual_headers = tuple(
                "" if cell.value is None else str(cell.value).strip()
                for cell in sheet[1][: len(HEADERS)]
            )
            if actual_headers != HEADERS:
                raise EquipmentError(
                    "Столбцы Excel не соответствуют шаблону equipment_data.xlsx"
                )
            rows = [
                tuple(cell.value for cell in row[: len(HEADERS)])
                for row in sheet.iter_rows(min_row=2)
                if any(not _is_empty(cell.value) for cell in row[: len(HEADERS)])
            ]
        finally:
            workbook.close()

        if not rows:
            raise EquipmentError("В Excel нет строк оборудования")

        equipment, errors = self._validate_rows(rows, substance_ids)
        if errors:
            report_path = project / ERROR_FILE_NAME
            report_path.write_text("\n".join(errors) + "\n", encoding="utf-8")
            raise EquipmentError(
                f"Импорт остановлен: найдено ошибок — {len(errors)}. "
                f"Протокол: {report_path}"
            )

        input_path = project / "input" / EXCEL_FILE_NAME
        input_path.parent.mkdir(parents=True, exist_ok=True)
        if source.resolve() != input_path.resolve():
            shutil.copy2(source, input_path)

        json_path = project / JSON_FILE_NAME
        temporary_path = project / f".{JSON_FILE_NAME}.tmp"
        temporary_path.write_text(
            json.dumps(equipment, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        temporary_path.replace(json_path)
        error_path = project / ERROR_FILE_NAME
        if error_path.exists():
            error_path.unlink()
        return EquipmentImportResult(len(equipment), json_path, input_path)

    def _validate_rows(
        self, rows: list[tuple[Any, ...]], substance_ids: set[int]
    ) -> tuple[list[dict[str, Any]], list[str]]:
        result: list[dict[str, Any]] = []
        errors: list[str] = []
        source_ids: set[int] = set()

        for project_id, values in enumerate(rows, start=1):
            row_number = project_id + 1
            row = dict(zip(HEADERS, values))
            source_id = _integer(
                row["source_id"], "source_id", errors, row_number, required=False
            )
            if source_id is not None:
                _positive(float(source_id), "source_id", errors, row_number)
                if source_id in source_ids:
                    errors.append(
                        f"строка {row_number}, source_id: значение {source_id} повторяется"
                    )
                source_ids.add(source_id)

            substance_id = _integer(
                row["substance_id"], "substance_id", errors, row_number
            )
            if substance_id is not None and substance_id not in substance_ids:
                errors.append(
                    f"строка {row_number}, substance_id: ID {substance_id} "
                    "отсутствует в substances.json"
                )
            equipment_name = _text(
                row["equipment_name"], "equipment_name", errors, row_number
            )
            equipment_type = _integer(
                row["equipment_type"], "equipment_type", errors, row_number
            )
            if equipment_type is not None and equipment_type not in range(10):
                errors.append(
                    f"строка {row_number}, equipment_type: ожидается код от 0 до 9"
                )
            phase_state = _text(
                row["phase_state"], "phase_state", errors, row_number
            )
            if phase_state and phase_state not in PHASE_STATES:
                errors.append(
                    f"строка {row_number}, phase_state: неизвестное значение {phase_state!r}"
                )

            total_length = _number(
                row["total_length_m"],
                "total_length_m",
                errors,
                row_number,
                required=equipment_type in PIPELINE_TYPES,
            )
            equipment_count = _integer(
                row["equipment_count"],
                "equipment_count",
                errors,
                row_number,
                required=equipment_type is not None and equipment_type not in PIPELINE_TYPES,
            )
            accident_length = _number(
                row["accident_section_length_m"],
                "accident_section_length_m",
                errors,
                row_number,
                required=equipment_type in PIPELINE_TYPES,
            )
            diameter = _number(
                row["diameter_mm"],
                "diameter_mm",
                errors,
                row_number,
                required=equipment_type in PIPELINE_TYPES,
            )
            wall = _number(
                row["wall_thickness_mm"],
                "wall_thickness_mm",
                errors,
                row_number,
                required=equipment_type in PIPELINE_TYPES,
            )
            volume = _number(
                row["volume_m3"],
                "volume_m3",
                errors,
                row_number,
                required=equipment_type is not None and equipment_type not in PIPELINE_TYPES,
            )

            if equipment_type in PIPELINE_TYPES:
                for value, field in (
                    (total_length, "total_length_m"),
                    (accident_length, "accident_section_length_m"),
                    (diameter, "diameter_mm"),
                    (wall, "wall_thickness_mm"),
                ):
                    _positive(value, field, errors, row_number)
                if (
                    total_length is not None
                    and accident_length is not None
                    and accident_length > total_length
                ):
                    errors.append(
                        f"строка {row_number}, accident_section_length_m: "
                        "аварийный участок больше полной длины"
                    )
                if diameter is not None and wall is not None and wall * 2 >= diameter:
                    errors.append(
                        f"строка {row_number}, wall_thickness_mm: "
                        "внутренний диаметр должен быть больше нуля"
                    )
                if equipment_count is not None:
                    errors.append(
                        f"строка {row_number}, equipment_count: для трубопровода поле не заполняется"
                    )
            elif equipment_type is not None:
                _positive(
                    float(equipment_count) if equipment_count is not None else None,
                    "equipment_count",
                    errors,
                    row_number,
                )
                _positive(volume, "volume_m3", errors, row_number)
                if total_length is not None or accident_length is not None:
                    errors.append(
                        f"строка {row_number}: total_length_m и accident_section_length_m "
                        "заполняются только для трубопроводов"
                    )

            fill_fraction = _number(
                row["fill_fraction"], "fill_fraction", errors, row_number
            )
            if fill_fraction is not None and not 0 < fill_fraction <= 1:
                errors.append(
                    f"строка {row_number}, fill_fraction: ожидается значение больше 0 и не больше 1"
                )
            pressure = _number(row["pressure_mpa"], "pressure_mpa", errors, row_number)
            spill_coefficient = _number(
                row["spill_coefficient"], "spill_coefficient", errors, row_number
            )
            spill_area = _number(
                row["spill_area_m2"], "spill_area_m2", errors, row_number, required=False
            )
            temperature = _number(
                row["substance_temperature_c"],
                "substance_temperature_c",
                errors,
                row_number,
            )
            shutdown_time = _number(
                row["shutdown_time_s"], "shutdown_time_s", errors, row_number
            )
            evaporation_time = _number(
                row["evaporation_time_s"], "evaporation_time_s", errors, row_number
            )
            for value, field in (
                (pressure, "pressure_mpa"),
                (spill_coefficient, "spill_coefficient"),
                (spill_area, "spill_area_m2"),
                (shutdown_time, "shutdown_time_s"),
                (evaporation_time, "evaporation_time_s"),
            ):
                _non_negative(value, field, errors, row_number)

            hazard_component = _text(
                row["hazard_component"], "hazard_component", errors, row_number
            )
            clutter_degree = _integer(
                row["clutter_degree"], "clutter_degree", errors, row_number
            )
            if clutter_degree is not None and clutter_degree not in range(1, 5):
                errors.append(
                    f"строка {row_number}, clutter_degree: ожидается код от 1 до 4"
                )
            coord_type = _integer(
                row["coord_type"], "coord_type", errors, row_number
            )
            _non_negative(coord_type, "coord_type", errors, row_number)
            coordinate_x = _number(
                row["coordinate_x"], "coordinate_x", errors, row_number
            )
            coordinate_y = _number(
                row["coordinate_y"], "coordinate_y", errors, row_number
            )
            possible_dead = _integer(
                row["possible_dead"], "possible_dead", errors, row_number
            )
            possible_injured = _integer(
                row["possible_injured"], "possible_injured", errors, row_number
            )
            _non_negative(possible_dead, "possible_dead", errors, row_number)
            _non_negative(possible_injured, "possible_injured", errors, row_number)

            result.append(
                {
                    "id": project_id,
                    "source_id": source_id,
                    "substance_id": substance_id,
                    "equipment_name": equipment_name,
                    "equipment_type": equipment_type,
                    "phase_state": phase_state,
                    "total_length_m": total_length,
                    "equipment_count": equipment_count,
                    "accident_section_length_m": accident_length,
                    "diameter_mm": diameter,
                    "wall_thickness_mm": wall,
                    "volume_m3": volume,
                    "fill_fraction": fill_fraction,
                    "pressure_mpa": pressure,
                    "spill_coefficient": spill_coefficient,
                    "spill_area_m2": 0.0 if spill_area is None else spill_area,
                    "substance_temperature_c": temperature,
                    "shutdown_time_s": shutdown_time,
                    "evaporation_time_s": evaporation_time,
                    "hazard_component": hazard_component,
                    "clutter_degree": clutter_degree,
                    "coord_type": coord_type,
                    "coordinates": [coordinate_x, coordinate_y],
                    "possible_dead": possible_dead,
                    "possible_injured": possible_injured,
                }
            )

        return result, errors
