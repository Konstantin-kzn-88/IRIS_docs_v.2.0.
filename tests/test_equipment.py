import json
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from iris_v2.equipment import HEADERS, EquipmentError, EquipmentService


def write_project(project: Path, substances: list[dict] | None = None) -> None:
    (project / "input").mkdir(parents=True)
    (project / "substances.json").write_text(
        json.dumps(
            substances or [{"id": 1, "name": "Нефть", "kind": 0}],
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )


def write_excel(path: Path, values: list) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Equipment Data"
    sheet.append(list(HEADERS))
    sheet.append(values)
    workbook.save(path)


def pipeline_row(substance_id: int = 1, accident_length: float = 100) -> list:
    return [
        17,
        substance_id,
        "Нефтепровод",
        9,
        "ж.ф.",
        1000,
        None,
        accident_length,
        219,
        8,
        None,
        0.8,
        1.6,
        20,
        0,
        20,
        12,
        3600,
        "Участок трубопроводов",
        2,
        0,
        "[10.5, 20.5, 30]",
        2,
        4,
    ]


def test_template_is_copied_to_project(tmp_path: Path) -> None:
    project = tmp_path / "project"
    write_project(project)

    template = EquipmentService().ensure_template(project)

    assert template == project / "input" / "equipment_data.xlsx"
    assert template.read_bytes().startswith(b"PK")
    workbook = load_workbook(template, read_only=True, data_only=True)
    try:
        rows = [
            row
            for row in workbook["Equipment Data"].iter_rows(
                min_row=2, max_col=len(HEADERS), values_only=True
            )
            if any(value is not None for value in row)
        ]
    finally:
        workbook.close()
    assert len(rows) == 9
    assert [row[3] for row in rows] == [0, 1, 2, 3, 4, 6, 7, 8, 9]
    assert all(row[1] == 1 for row in rows)
    assert all(row[4] == "ж.ф." for row in rows)
    result = EquipmentService().import_excel(project, template)
    imported = json.loads(result.json_path.read_text(encoding="utf-8"))
    assert result.count == 9
    assert [item["equipment_type"] for item in imported] == [
        0, 1, 2, 3, 4, 6, 7, 8, 9
    ]


def test_template_uses_every_selected_substance(tmp_path: Path) -> None:
    project = tmp_path / "project"
    write_project(
        project,
        [
            {"id": 1, "name": "Нефть", "kind": 0},
            {"id": 2, "name": "Бензин", "kind": 0},
            {"id": 3, "name": "Газ", "kind": 2},
        ],
    )

    template = EquipmentService().ensure_template(project)
    workbook = load_workbook(template, read_only=True, data_only=True)
    try:
        rows = [
            row
            for row in workbook["Equipment Data"].iter_rows(
                min_row=2, max_col=len(HEADERS), values_only=True
            )
            if any(value is not None for value in row)
        ]
    finally:
        workbook.close()

    assert len(rows) == 24
    assert [row[3] for row in rows if row[1] == 3] == [0, 2, 3, 5, 6, 8]
    assert all(row[4] == "г.ф." for row in rows if row[1] == 3)
    assert "Нефть" in rows[0][2]
    assert "Бензин" in rows[9][2]


def test_pipeline_excel_is_imported(tmp_path: Path) -> None:
    project = tmp_path / "project"
    write_project(project)
    source = tmp_path / "source.xlsx"
    write_excel(source, pipeline_row())

    result = EquipmentService().import_excel(project, source)
    equipment = json.loads(result.json_path.read_text(encoding="utf-8"))

    assert result.count == 1
    assert result.excel_path == project / "input" / "equipment_data.xlsx"
    assert equipment[0]["id"] == 1
    assert equipment[0]["source_id"] == 17
    assert equipment[0]["total_length_m"] == 1000.0
    assert equipment[0]["accident_section_length_m"] == 100.0
    assert equipment[0]["equipment_count"] is None
    assert equipment[0]["coordinates"] == [10.5, 20.5, 30.0]


def test_errors_block_import_and_create_report(tmp_path: Path) -> None:
    project = tmp_path / "project"
    write_project(project)
    source = tmp_path / "source.xlsx"
    row = pipeline_row(substance_id=999, accident_length=1100)
    row[12] = 0
    row[21] = "0, 0"
    write_excel(source, row)

    with pytest.raises(EquipmentError, match="Импорт остановлен"):
        EquipmentService().import_excel(project, source)

    assert not (project / "equipments.json").exists()
    report = (project / "equipment_import_errors.txt").read_text(encoding="utf-8")
    assert "substance_id: ID 999 отсутствует" in report
    assert "аварийный участок больше полной длины" in report
    assert "pressure_mpa: значение должно быть не меньше 0,1" in report
    assert "coordinates: ожидается список чисел" in report
